"""
Write GATI-format Excel outputs using sheettoadd.xlsx.

CSD holds the OPS form with formulas that reference row 2 of Gati.
Gati holds the processed order data (standard GATI columns).
"""

from __future__ import annotations

import os

import openpyxl
import pandas as pd

TEMPLATE_FILENAME = "sheettoadd.xlsx"
CSD_SHEET_NAME = "CSD"
DATA_SHEET_NAME = "Gati"


def template_path() -> str:
    return os.path.join(os.path.dirname(os.path.abspath(__file__)), TEMPLATE_FILENAME)


def _canonical_headers(ws) -> list[str]:
    headers: list[str] = []
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col).value
        if val is None or str(val).strip() == "":
            break
        headers.append(str(val).strip())
    return headers


def _align_dataframe(df: pd.DataFrame, canonical: list[str]) -> pd.DataFrame:
    aligned = pd.DataFrame()
    for col in canonical:
        aligned[col] = df[col] if col in df.columns else ""
    return aligned.fillna("")


def write_gati_excel(df: pd.DataFrame, output_path: str, template: str | None = None, csd_remarks: str | None = None) -> str:
    """
    Save processed data to Gati of the template workbook, preserving CSD formulas.
    Optionally set remarks in CSD sheet.
    """
    tpl = template or template_path()
    if not os.path.exists(tpl):
        raise FileNotFoundError(f"GATI template not found: {tpl}")

    wb = openpyxl.load_workbook(tpl)
    ws2 = wb[DATA_SHEET_NAME]
    canonical = _canonical_headers(ws2)
    if not canonical:
        raise ValueError(f"No headers found on {DATA_SHEET_NAME} in {tpl}")

    df_out = _align_dataframe(df, canonical)

    if ws2.max_row > 1:
        ws2.delete_rows(2, ws2.max_row - 1)

    for row_idx, row in enumerate(df_out.itertuples(index=False, name=None), start=2):
        for col_idx, value in enumerate(row, start=1):
            cell_value = value
            if pd.isna(cell_value):
                cell_value = ""
            ws2.cell(row=row_idx, column=col_idx, value=cell_value)

    ws2.sheet_state = "hidden"
    if CSD_SHEET_NAME in wb.sheetnames:
        wb[CSD_SHEET_NAME].sheet_state = "visible"
        wb.active = wb[CSD_SHEET_NAME]
        
        # If csd_remarks is provided, set row 22 column 3 of CSD sheet
        if csd_remarks is not None:
            wb[CSD_SHEET_NAME].cell(row=22, column=3, value=csd_remarks)

    wb.save(output_path)
    return output_path


def enhance_xlsx_with_template(output_path: str, template: str | None = None, csd_remarks: str | None = None) -> str:
    """
    Replace a single-sheet processor output with the GATI template layout.
    Optionally set remarks in CSD sheet.
    """
    df = pd.read_excel(output_path)
    return write_gati_excel(df, output_path, template=template, csd_remarks=csd_remarks)
