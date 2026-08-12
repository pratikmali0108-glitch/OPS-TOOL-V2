import re
import os
import pandas as pd
import pdfplumber
from datetime import datetime
from openpyxl import load_workbook


REFERENCE_EXCEL_PATH = os.path.join(
    os.path.dirname(os.path.abspath(__file__)),
    "CS_100826",
    "DHI_CS_100826.xlsx"
)

OUTPUT_TEMPLATE_PATH = os.path.join(
    os.path.dirname(os.path.abspath(__file__)),
    "CS_100826",
    "DHI_CS_100826.xlsx"
)


TONE_MAP = {
    "YG": {
        "metal_kt": "Y",
        "tone": "Y",
        "metal_code": "G14Y",
        "metal_name": "YELLOW GOLD"
    },
    "WG": {
        "metal_kt": "W",
        "tone": "W",
        "metal_code": "G14W",
        "metal_name": "WHITE GOLD"
    },
    "PG": {
        "metal_kt": "P",
        "tone": "P",
        "metal_code": "G14P",
        "metal_name": "PINK GOLD"
    },
    "PT": {
        "metal_kt": "PT",
        "tone": "PT",
        "metal_code": "PT",
        "metal_name": "PLATINUM"
    },
    "AL": {
        "metal_kt": "AL",
        "tone": "AL",
        "metal_code": "AL",
        "metal_name": "ALLOY"
    },
}


def parse_metal_from_description(desc_text, recycled=False):
    desc_upper = str(desc_text).upper()

    metal_info = {
        "karat": "14",
        "tone_suffix": "YG",
        "metal": "G14Y",
        "tone": "Y",
        "metal_name": "YELLOW GOLD",
    }

    # ---------------------------------------------------------
    # Detect karat
    # ---------------------------------------------------------
    if re.search(r"\b18K\b", desc_upper):
        metal_info["karat"] = "18"
    elif re.search(r"\b10K\b", desc_upper):
        metal_info["karat"] = "10"
    elif re.search(r"\b22K\b", desc_upper):
        metal_info["karat"] = "22"
    elif re.search(r"\b14K\b", desc_upper):
        metal_info["karat"] = "14"

    # ---------------------------------------------------------
    # Detect metal/tone
    # IMPORTANT:
    # PT is checked FIRST.
    # ---------------------------------------------------------

    if (
        re.search(r"\bPLATINUM\b", desc_upper)
        or re.search(r"\bPLAT\b", desc_upper)
        or re.search(r"\bPT95\b", desc_upper)
        or re.search(r"\bPT\b", desc_upper)
    ):
        tone_suffix = "PT"

    elif (
        re.search(r"\bWHITE\s+GOLD\b", desc_upper)
        or re.search(r"\bWHITE\b", desc_upper)
        or re.search(r"\b14KW\b", desc_upper)
        or re.search(r"\b18KW\b", desc_upper)
        or re.search(r"\b10KW\b", desc_upper)
    ):
        tone_suffix = "WG"

    elif (
        re.search(r"\bPINK\s+GOLD\b", desc_upper)
        or re.search(r"\bROSE\s+GOLD\b", desc_upper)
        or re.search(r"\bPINK\b", desc_upper)
        or re.search(r"\bROSE\b", desc_upper)
        or re.search(r"\b14KP\b", desc_upper)
        or re.search(r"\b18KP\b", desc_upper)
    ):
        tone_suffix = "PG"

    elif (
        re.search(r"\bALLOY\b", desc_upper)
        or re.search(r"\b14KAL\b", desc_upper)
    ):
        tone_suffix = "AL"

    elif (
        re.search(r"\bYELLOW\s+GOLD\b", desc_upper)
        or re.search(r"\bYELLOW\b", desc_upper)
        or re.search(r"\b14KY\b", desc_upper)
        or re.search(r"\b18KY\b", desc_upper)
        or re.search(r"\b10KY\b", desc_upper)
    ):
        tone_suffix = "YG"

    else:
        tone_suffix = "YG"

    # ---------------------------------------------------------
    # Build final metal information
    # ---------------------------------------------------------
    metal_info["tone_suffix"] = tone_suffix

    t = TONE_MAP[tone_suffix]

    metal_info["tone"] = t["tone"]
    metal_info["metal_name"] = t["metal_name"]

    suffix_z = "Z" if recycled else ""

    if tone_suffix == "PT":
        metal_info["metal"] = f"PC95{suffix_z}"

    elif tone_suffix == "AL":
        metal_info["metal"] = f"AL{metal_info['karat']}{suffix_z}"

    else:
        metal_info["metal"] = (
            f"G{metal_info['karat']}{t['metal_kt']}{suffix_z}"
        )

    return metal_info

def extract_size_from_text(text):
    size = None

    patterns = [
        r"SZ\s*([\d.]+)",
        r"SIZE\s*([\d.]+)",
        r"(\d+(?:\.\d+)?)\s*[-–]\s*(?:YG|WG|PG|PT|AL|YGSM|PTM|AG)",
    ]

    for pattern in patterns:
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            size = float(match.group(1))
            if size == int(size):
                size = int(size)
            return str(size)

    return size


def format_itemsize_prefix(base_style, size_str, reference_df):
    matches = reference_df[reference_df["Style No"] == base_style]
    if not matches.empty:
        valid_items = matches["ItemSize"].dropna()
        if not valid_items.empty:
            first_itemsize = valid_items.iloc[0]
            prefix = first_itemsize.split()[0] if " " in first_itemsize else re.match(r"^([A-Za-z]+)", str(first_itemsize)).group(1)
        else:
            prefix = "TS"
    else:
        prefix = "TS"

    try:
        size_val = float(size_str)
        if size_val == int(size_val):
            size_num = int(size_val)
            if size_num < 10:
                return f"{prefix} {size_num:02d}"
            else:
                return f"{prefix} {size_num}"
        else:
            return f"{prefix} {size_str}"
    except ValueError:
        return f"{prefix} {size_str}"


def build_style_code(base_style, size_str, tone_suffix, reference_df, order_group=""):
    is_costco = (order_group or "").strip().lower() == "costco"

    candidate = f"{base_style}-{size_str}{tone_suffix}"
    if is_costco:
        candidate = f"{candidate}CO"

    lookup_candidate = candidate
    if is_costco:
        lookup_candidate = f"{base_style}-{size_str}{tone_suffix}"

    matches = reference_df[reference_df["Client Style No"] == lookup_candidate]
    if not matches.empty:
        row = matches.iloc[0]
        itemsize = row["ItemSize"] if pd.notna(row["ItemSize"]) else None
        return candidate, itemsize

    style_matches = reference_df[reference_df["Style No"] == base_style]
    if not style_matches.empty:
        for _, row in style_matches.iterrows():
            client_style = row["Client Style No"]
            if pd.notna(client_style):
                parts = str(client_style).rsplit("-", 1)
                if len(parts) == 2:
                    ref_tone_part = parts[1]
                    match = re.match(r"^([\d.]+)([A-Z]+)$", ref_tone_part, re.IGNORECASE)
                    if match and match.group(2).upper() == tone_suffix.upper():
                        itemsize = format_itemsize_prefix(base_style, size_str, reference_df)
                        return candidate, itemsize

    itemsize = format_itemsize_prefix(base_style, size_str, reference_df)
    return candidate, itemsize

def is_item_row(line):
    """
    Detect whether a line starts a new order/item row.

    Examples:
        1 SZ5 RG0003054K ...
        2 R63738-PT-EL-CSTRG0003054K ...
        3 SZ6 RG0003054K ...
        4 SZ8 RG0003054K ...
    """

    if not line:
        return False

    line = line.strip()

    pattern_a = re.match(
        r"^\s*\d+\s+SZ\s*[\d.]+\s+[A-Z]{2}\d{4,}[A-Z0-9]*\b",
        line,
        re.IGNORECASE
    )
    if pattern_a:
        return True

    pattern_b = re.match(
        r"^\s*\d+\s+\S+[A-Z]{2}\d{4,}[A-Z0-9]*\b",
        line,
        re.IGNORECASE
    )
    if pattern_b:
        style_part = re.search(r"[A-Z]{2}\d{4,}[A-Z0-9]*\b", line, re.IGNORECASE)
        if style_part and "$" in line:
            return True

    return False

def _extract_common_po_fields(full_text):
    """Extract PO-level fields shared by different VPO layouts."""
    def find(pattern):
        m = re.search(pattern, full_text, re.IGNORECASE | re.MULTILINE)
        return m.group(1).strip() if m else None

    order_number = find(r"Order\s*#:\s*(\d+)")
    po_number = find(r"P\.?O\.?\s*#:\s*([^\n]+)")
    po_date = find(r"\bDate:\s*(\d{1,2}/\d{1,2}/\d{2,4})")
    due_date = find(r"\bDue\s*Date:\s*(\d{1,2}/\d{1,2}/\d{2,4})")

    # Some VPO layouts put PO information on the same line/without labels.
    if po_number:
        po_number = po_number.strip()

    dia_quality = ""
    dq_match = re.search(r"Diamond\s+Quality\s+([^\n]+)", full_text, re.IGNORECASE)
    if dq_match:
        dia_quality = dq_match.group(1).strip()

    stamping = ""
    stamp_match = re.search(r"Stamping\s+([^\n]+)", full_text, re.IGNORECASE)
    if stamp_match:
        stamping = stamp_match.group(1).strip()

    return {
        "order_number": order_number,
        "po_number": po_number,
        "po_date": po_date,
        "due_date": due_date,
        "dia_quality": dia_quality,
        "stamping": stamping,
    }


def _clean_pdf_lines(full_text):
    """Normalize pdfplumber output without destroying meaningful item text."""
    result = []

    for raw_line in full_text.splitlines():
        line = re.sub(r"\s+", " ", raw_line).strip()

        if not line:
            result.append("")
            continue

        # Remove the common copyright/page noise only when it is clearly noise.
        if re.match(r"^(RightClick®|Page\s*#?:|Grand\s+Total:?)", line, re.IGNORECASE):
            continue

        result.append(line)

    return result


def _parse_standard_vpo_layout(lines, common):
    """
    Existing-style parser.

    This retains the original parser's recognition strategy but makes the
    header detection less brittle and isolates it from the alternate layout.
    """
    orders = []

    header_idx = None
    for idx, line in enumerate(lines):
        if re.search(
            r"Memo.*Item.*Description.*Size.*Quantity",
            line,
            re.IGNORECASE,
        ):
            header_idx = idx
            break

    if header_idx is None:
        # Try the common header in a more tolerant order.
        for idx, line in enumerate(lines):
            if (
                re.search(r"\bMemo\b", line, re.IGNORECASE)
                and re.search(r"\bItem\b", line, re.IGNORECASE)
                and re.search(r"\bDescription\b", line, re.IGNORECASE)
            ):
                header_idx = idx
                break

    if header_idx is None:
        return orders

    memo_prefix = ""
    i = header_idx + 1

    def _is_new_memo_prefix_line(cl):
        ct = cl.split()
        return bool(
            len(ct) >= 1
            and (
                re.match(r"^[A-Z]\d{4,}[-]", ct[0], re.IGNORECASE)
                or re.match(r"^SKU#", ct[0], re.IGNORECASE)
            )
        )

    while i < len(lines):
        line = lines[i].strip()

        if not line:
            i += 1
            continue

        if (
            re.match(r"^Diamond\s+Quality", line, re.IGNORECASE)
            or re.match(r"^Stamping", line, re.IGNORECASE)
            or re.match(r"^Grand\s+Total", line, re.IGNORECASE)
            or re.match(r"^RightClick", line, re.IGNORECASE)
            or re.match(r"^\d{4,}", line)
        ):
            i += 1
            continue

        line_tokens = line.split()
        memo_prefix_line_captured = None

        if memo_prefix == "" and _is_new_memo_prefix_line(line):
            memo_prefix_line_captured = line
            memo_prefix = line_tokens[0]
            desc_from_prefix = " ".join(line_tokens[1:]) if len(line_tokens) > 1 else ""
            i += 1

            if i < len(lines):
                next_line = lines[i].strip()
                next_tokens = next_line.split()

                item_patt = re.compile(r"[A-Z]{2}\d{4,}[A-Z0-9]*")
                style_idx = None
                base_style = None

                for ti, tok in enumerate(next_tokens):
                    m = item_patt.search(tok)
                    if m:
                        style_idx = ti
                        base_style = m.group()
                        break

                if style_idx is not None and style_idx >= 1:
                    memo_suffix = "".join(next_tokens[1:style_idx])
                    full_memo = (memo_prefix + memo_suffix).replace(" ", "")

                    price_idx = None
                    for ti in range(len(next_tokens) - 1, -1, -1):
                        if "$" in next_tokens[ti]:
                            price_idx = ti
                            break

                    size_val = None
                    qty_val = 1

                    if price_idx is not None and price_idx >= 4:
                        weight_idx = price_idx - 2
                        qty_idx = weight_idx - 1
                        size_idx = weight_idx - 2

                        if size_idx > style_idx:
                            size_val = next_tokens[size_idx]

                        if qty_idx > style_idx:
                            try:
                                qty_val = int(next_tokens[qty_idx])
                            except (ValueError, TypeError):
                                qty_val = 1

                    sz_match = re.search(r"SZ\s*([\d.]+)", next_line, re.IGNORECASE)
                    if sz_match and size_val is None:
                        size_val = sz_match.group(1)

                    desc_parts = []
                    if desc_from_prefix:
                        desc_parts.append(desc_from_prefix)

                    if price_idx is not None:
                        slice_end = (
                            price_idx - 5
                            if (price_idx - 5) > (style_idx + 1)
                            else len(next_tokens)
                        )
                        mid_tokens = next_tokens[style_idx + 1:slice_end]
                    else:
                        mid_tokens = next_tokens[style_idx + 1:]

                    filtered_mid = []
                    for ti, tok in enumerate(mid_tokens):
                        if tok == "$" or tok.startswith("$"):
                            break
                        if tok == "Q" and ti >= len(mid_tokens) - 3:
                            break
                        filtered_mid.append(tok)

                    if filtered_mid:
                        desc_parts.append(" ".join(filtered_mid))

                    j = i + 1
                    while j < len(lines) and j < i + 4:
                        cont_line = lines[j].strip()

                        if is_item_row(cont_line) or _is_new_memo_prefix_line(cont_line):
                            break

                        cont_tokens = cont_line.split()

                        is_summary_line = (
                            len(cont_tokens) == 2
                            and re.match(r"^\d+$", cont_tokens[0])
                            and re.match(r"^[\d.]+$", cont_tokens[1])
                        )

                        is_meta_line = bool(
                            re.match(
                                r"^(Diamond|Stamping|Grand\s+Total|RightClick|\d{8})",
                                cont_line,
                                re.IGNORECASE,
                            )
                        )

                        if is_summary_line or is_meta_line or not cont_line:
                            break

                        desc_parts.append(cont_line)
                        j += 1

                    full_description = " ".join(desc_parts)
                    full_description = re.sub(r"\s{2,}", " ", full_description).strip()

                    size = None
                    sz_match = re.search(
                        r"SZ\s*([\d.]+)",
                        full_description,
                        re.IGNORECASE,
                    )
                    if sz_match:
                        size = sz_match.group(1)
                    elif size_val is not None:
                        size = size_val

                    if size:
                        try:
                            sv = float(size)
                            if sv == int(sv):
                                size = str(int(sv))
                        except (ValueError, TypeError):
                            pass

                    cleaned_desc = re.sub(
                        r"\b\d+\s+\d\s+[\d.]+\b",
                        "",
                        full_description,
                    )
                    cleaned_desc = re.sub(r"\s{2,}", " ", cleaned_desc).strip()

                    raw_block = (
                        (memo_prefix_line_captured or "") + " " + next_line
                    ).strip()

                    orders.append({
                        "base_style": base_style,
                        "description": cleaned_desc,
                        "raw_description_block": raw_block,
                        "size": size if size else "1",
                        "qty": qty_val,
                        "memo": full_memo,
                        "dia_quality": common["dia_quality"],
                        "stamping": common["stamping"],
                        "order_number": common["order_number"],
                        "po_number": common["po_number"],
                        "po_date": common["po_date"],
                        "due_date": common["due_date"],
                    })

                    memo_prefix = ""
                    i = j
                    continue

            memo_prefix = ""
            i += 1
            continue

        item_patt = re.compile(r"[A-Z]{2}\d{4,}[A-Z0-9]*")
        item_match = item_patt.search(line)

        if item_match and "$" in line:
            base_style = item_match.group()
            next_tokens = line_tokens

            style_idx = None
            for ti, tok in enumerate(next_tokens):
                if item_patt.search(tok):
                    style_idx = ti
                    break

            memo_full = ""
            if style_idx is not None and style_idx >= 2:
                memo_parts = next_tokens[1:style_idx]
                memo_full = (
                    (memo_prefix + "".join(memo_parts)).replace(" ", "")
                    if memo_prefix
                    else "".join(memo_parts)
                )

            price_idx = None
            for ti in range(len(next_tokens) - 1, -1, -1):
                if "$" in next_tokens[ti]:
                    price_idx = ti
                    break

            qty_val = 1
            size_val = None

            if price_idx is not None and price_idx >= 4:
                try:
                    qty_val = int(next_tokens[price_idx - 3])
                except (ValueError, IndexError, TypeError):
                    qty_val = 1

                try:
                    size_val = next_tokens[price_idx - 4]
                except IndexError:
                    size_val = None

            sz_match = re.search(r"SZ\s*([\d.]+)", line, re.IGNORECASE)
            if sz_match and size_val is None:
                size_val = sz_match.group(1)

            desc_parts = []
            if style_idx is not None and price_idx is not None:
                slice_end = (
                    price_idx - 5
                    if (price_idx - 5) > (style_idx + 1)
                    else len(next_tokens)
                )
                desc_parts.append(
                    " ".join(next_tokens[style_idx + 1:slice_end])
                )

            j = i + 1
            while j < len(lines) and j < i + 4:
                cont_line = lines[j].strip()

                if is_item_row(cont_line) or _is_new_memo_prefix_line(cont_line):
                    break

                cont_tokens = cont_line.split()

                is_summary_line = (
                    len(cont_tokens) == 2
                    and re.match(r"^\d+$", cont_tokens[0])
                    and re.match(r"^[\d.]+$", cont_tokens[1])
                )

                is_meta_line = bool(
                    re.match(
                        r"^(Diamond|Stamping|Grand\s+Total|RightClick|\d{8})",
                        cont_line,
                        re.IGNORECASE,
                    )
                )

                if is_summary_line or is_meta_line or not cont_line:
                    break

                desc_parts.append(cont_line)
                j += 1

            full_description = " ".join(desc_parts)
            full_description = re.sub(r"\s{2,}", " ", full_description).strip()

            size = None
            sz_match = re.search(
                r"SZ\s*([\d.]+)",
                full_description,
                re.IGNORECASE,
            )
            if sz_match:
                size = sz_match.group(1)
            elif size_val is not None:
                size = size_val

            if size:
                try:
                    sv = float(size)
                    if sv == int(sv):
                        size = str(int(sv))
                except (ValueError, TypeError):
                    pass

            raw_block_full = line
            if memo_prefix_line_captured:
                raw_block_full = memo_prefix_line_captured + " " + raw_block_full

            orders.append({
                "base_style": base_style,
                "description": full_description,
                "raw_description_block": raw_block_full,
                "size": size if size else "1",
                "qty": qty_val,
                "memo": memo_full,
                "dia_quality": common["dia_quality"],
                "stamping": common["stamping"],
                "order_number": common["order_number"],
                "po_number": common["po_number"],
                "po_date": common["po_date"],
                "due_date": common["due_date"],
            })

            memo_prefix = ""
            i = j
            continue

        i += 1

    return orders


def _parse_alternate_vpo_layout(lines, common):
    """
    Parser for VPO layouts containing:

        Item #
        Quantity
        Vendor Item #
        Cost
        Amount
        ...
        Description
        Memo #

    Example item block:

        1
        108
        $0.00
        $0.00
        SKU:1507287 14KW
        1.40CTW I VS Round
        Diamond Studs w/
        Gurdian Back (Only for labor)
        0.0000
        Q
        ER140-14KW-SEMI
        108
        0.0000
    """
    orders = []

    # Locate the alternate header.
    header_idx = None
    for idx, line in enumerate(lines):
        if (
            re.search(r"\bItem\s*#?\b", line, re.IGNORECASE)
            and re.search(r"\bQuantity\b", line, re.IGNORECASE)
            and re.search(r"\bVendor\s+Item\s*#?\b", line, re.IGNORECASE)
            and re.search(r"\bDescription\b", line, re.IGNORECASE)
        ):
            header_idx = idx
            break

    # In pdfplumber the header can be split into separate lines.
    if header_idx is None:
        item_idx = next(
            (
                i for i, x in enumerate(lines)
                if re.fullmatch(r"#?\s*Item\s*#?", x, re.IGNORECASE)
            ),
            None,
        )
        if item_idx is not None:
            window = " ".join(lines[item_idx:item_idx + 20])
            if (
                re.search(r"\bQuantity\b", window, re.IGNORECASE)
                and re.search(r"\bVendor\s+Item\b", window, re.IGNORECASE)
                and re.search(r"\bDescription\b", window, re.IGNORECASE)
            ):
                header_idx = item_idx

    if header_idx is None:
        return orders

    # We need a numbered item followed somewhere by a quantity.
    i = header_idx + 1

    while i < len(lines):
        line = lines[i].strip()

        # Skip obvious page/summary/footer content.
        if not line:
            i += 1
            continue

        if re.match(
            r"^(Purchase Order|DHARM INTERNATIONAL LLC|Grand\s+Total|"
            r"RightClick|Page\s*#?:)",
            line,
            re.IGNORECASE,
        ):
            i += 1
            continue

        # Item number is a standalone integer.
        if not re.fullmatch(r"\d+", line):
            i += 1
            continue

        item_no = line

        # Quantity should follow the item number.
        if i + 1 >= len(lines) or not re.fullmatch(r"\d+", lines[i + 1].strip()):
            i += 1
            continue

        qty_val = int(lines[i + 1].strip())
        j = i + 2

        # Skip cost and amount.
        money_seen = 0
        while j < len(lines) and money_seen < 2:
            token = lines[j].strip()

            if re.fullmatch(r"\$[\d,]+(?:\.\d+)?", token):
                money_seen += 1
                j += 1
                continue

            # Be tolerant of OCR/pdfplumber splitting "$" and amount.
            if token == "$" and j + 1 < len(lines):
                if re.fullmatch(r"[\d,]+(?:\.\d+)?", lines[j + 1].strip()):
                    money_seen += 1
                    j += 2
                    continue

            break

        # Search for the description/memo area until the next item/footer.
        block_start = j
        block = []

        while j < len(lines):
            token = lines[j].strip()

            if not token:
                j += 1
                continue

            # A new numbered item with a following quantity starts a new row.
            if (
                re.fullmatch(r"\d+", token)
                and j + 1 < len(lines)
                and re.fullmatch(r"\d+", lines[j + 1].strip())
            ):
                break

            if re.match(
                r"^(All findings|Purchase Order|DHARM INTERNATIONAL LLC|"
                r"Grand\s+Total|RightClick|Page\s*#?:)",
                token,
                re.IGNORECASE,
            ):
                break

            block.append(token)
            j += 1

        if not block:
            i += 1
            continue

        # Vendor item/style is generally an alphanumeric code containing
        # both letters and digits, often with hyphens.
        style_candidates = []

        for pos, token in enumerate(block):
            clean = token.strip()

            # Exclude obvious non-style numeric values.
            if re.fullmatch(r"[\d.]+", clean):
                continue

            # Strong style-code candidates.
            if (
                re.search(r"[A-Z]", clean, re.IGNORECASE)
                and re.search(r"\d", clean)
                and (
                    "-" in clean
                    or re.fullmatch(r"[A-Z]{1,6}\d[A-Z0-9\-]*", clean, re.IGNORECASE)
                )
            ):
                style_candidates.append((pos, clean))

        # Prefer the candidate appearing later in the block. In this layout
        # Vendor Item # occurs after weight/unit/description.
        base_style = None
        style_pos = None

        if style_candidates:
            style_pos, base_style = style_candidates[-1]

        # SKU/memo extraction.
        memo = ""
        sku_match = None

        for token in block:
            m = re.search(r"\bSKU\s*[:#]?\s*([A-Z0-9._/\-]+)", token, re.IGNORECASE)
            if m:
                sku_match = m.group(1)
                break

        if sku_match:
            memo = sku_match

        # Description is everything before the trailing weight/unit/vendor
        # item section. Keep the metal token such as 14KW in the description
        # because parse_metal_from_description() uses it.
        desc_tokens = []

        if style_pos is not None:
            pre_style = block[:style_pos]
        else:
            pre_style = block

        # Remove obvious trailing numeric fields while retaining real
        # description text and SKU/metal information.
        for token in pre_style:
            if re.fullmatch(r"[\d.]+", token):
                continue
            if re.fullmatch(r"[A-Z]", token) and token.upper() in {"Q", "EA", "PCS"}:
                continue
            desc_tokens.append(token)

        description = " ".join(desc_tokens)
        description = re.sub(r"\s+", " ", description).strip()

        # If SKU exists, preserve the SKU line because it may contain the
        # metal/tone (e.g. "SKU:1507287 14KW").
        if not description:
            description = " ".join(block)

        # Extract size from explicit SZ/SIZE or from common "1.40CTW"
        # descriptions only if it actually looks like a ring size.
        size = extract_size_from_text(description)

        # For this layout there is often no item size. Do NOT incorrectly
        # use diamond carat weight as the size.
        if size is None:
            size = "1"

        raw_block = " ".join(block)

        orders.append({
            "base_style": base_style or "",
            "description": description,
            "raw_description_block": raw_block,
            "size": size,
            "qty": qty_val,
            "memo": memo,
            "dia_quality": common["dia_quality"],
            "stamping": common["stamping"],
            "order_number": common["order_number"],
            "po_number": common["po_number"],
            "po_date": common["po_date"],
            "due_date": common["due_date"],
        })

        i = j

    return orders


def _fallback_generic_item_parser(lines, common):
    """
    Last-resort parser.

    This prevents a valid VPO from being rejected solely because its
    presentation/order of columns changed. It only accepts a row when it can
    identify both a numeric quantity and a plausible vendor/style code.
    """
    orders = []

    for i, line in enumerate(lines):
        if not re.fullmatch(r"\d+", line):
            continue

        # Quantity immediately after item number.
        if i + 1 >= len(lines) or not re.fullmatch(r"\d+", lines[i + 1]):
            continue

        qty = int(lines[i + 1])

        block = []
        j = i + 2

        while j < len(lines) and len(block) < 20:
            token = lines[j].strip()

            if (
                re.fullmatch(r"\d+", token)
                and j + 1 < len(lines)
                and re.fullmatch(r"\d+", lines[j + 1].strip())
            ):
                break

            if re.match(
                r"^(Purchase Order|Grand\s+Total|RightClick|Page\s*#?)",
                token,
                re.IGNORECASE,
            ):
                break

            if token:
                block.append(token)

            j += 1

        # Find a plausible alphanumeric/hyphenated vendor style.
        style = None
        style_pos = None

        for pos, token in enumerate(block):
            if (
                re.search(r"[A-Z]", token, re.IGNORECASE)
                and re.search(r"\d", token)
                and (
                    "-" in token
                    or re.fullmatch(r"[A-Z]{1,8}\d[A-Z0-9]*", token, re.IGNORECASE)
                )
            ):
                # Prefer codes with a hyphen, since those are common Vendor
                # Item # values.
                if "-" in token:
                    style = token
                    style_pos = pos

        if style is None:
            for pos, token in enumerate(block):
                if (
                    re.search(r"[A-Z]", token, re.IGNORECASE)
                    and re.search(r"\d", token)
                ):
                    style = token
                    style_pos = pos

        if style is None:
            continue

        memo = ""
        for token in block:
            m = re.search(
                r"\bSKU\s*[:#]?\s*([A-Z0-9._/\-]+)",
                token,
                re.IGNORECASE,
            )
            if m:
                memo = m.group(1)
                break

        description_tokens = block[:style_pos] if style_pos is not None else block
        description = " ".join(description_tokens)
        description = re.sub(r"\s+", " ", description).strip()

        # Do not turn diamond weights into item sizes.
        size = extract_size_from_text(description) or "1"

        orders.append({
            "base_style": style,
            "description": description,
            "raw_description_block": " ".join(block),
            "size": size,
            "qty": qty,
            "memo": memo,
            "dia_quality": common["dia_quality"],
            "stamping": common["stamping"],
            "order_number": common["order_number"],
            "po_number": common["po_number"],
            "po_date": common["po_date"],
            "due_date": common["due_date"],
        })

    return orders


def parse_pdf(pdf_path):
    """
    Robust VPO parser supporting multiple PDF layouts.

    Strategy:
      1. Extract all text with pdfplumber.
      2. Extract PO-level fields independently.
      3. Try the original/standard VPO layout.
      4. Try the alternate "Item # / Quantity / Vendor Item #" layout.
      5. Use a conservative generic fallback.
      6. Return an empty list only when no credible item row exists.
    """
    with pdfplumber.open(pdf_path) as pdf:
        full_text_parts = []

        for page in pdf.pages:
            text = page.extract_text(x_tolerance=2, y_tolerance=3)
            if text:
                full_text_parts.append(text)

    full_text = "\n".join(full_text_parts)

    if not full_text.strip():
        return []

    common = _extract_common_po_fields(full_text)
    lines = _clean_pdf_lines(full_text)

    # First: original layout.
    orders = _parse_standard_vpo_layout(lines, common)
    if orders:
        return orders

    # Second: alternate VPO layout from PDFs such as:
    # Item # / Quantity / Vendor Item # / Cost / Amount / Description / Memo #
    orders = _parse_alternate_vpo_layout(lines, common)
    if orders:
        return orders

    # Third: conservative fallback.
    orders = _fallback_generic_item_parser(lines, common)
    return orders

def load_reference_style_map():
    try:
        df = pd.read_excel(REFERENCE_EXCEL_PATH)
        df.columns = [str(c).strip() for c in df.columns]
        return df
    except Exception:
        return pd.DataFrame()


def generate_output_rows(orders, reference_df, recycled=False, order_group="", priority=""):
    output_rows = []
    order_group_val = (order_group or "").strip()

    for idx, order in enumerate(orders, start=1):
        metal_info = parse_metal_from_description(
            order["description"] + " " + order.get("raw_description_block", ""),
            recycled=recycled,
        )

        style_code, itemsize = build_style_code(
            order["base_style"],
            order["size"],
            metal_info["tone_suffix"],
            reference_df,
            order_group=order_group_val,
        )

        if itemsize is None:
            itemsize = format_itemsize_prefix(order["base_style"], order["size"], reference_df)

        metal_full_name = f"{metal_info['karat']} {metal_info['metal_name']}"

        special_remarks_parts = []
        if order_group_val:
            special_remarks_parts.append(order_group_val)
        if order["memo"]:
            special_remarks_parts.append(order["memo"])
        special_remarks_parts.append(metal_full_name)
        if order["dia_quality"]:
            dq_clean = order["dia_quality"].strip()
            dq_tokens = dq_clean.split()
            if len(dq_tokens) >= 2:
                dq_formatted = "-".join(dq_tokens[:2])
                if len(dq_tokens) > 2:
                    dq_formatted += " " + " ".join(dq_tokens[2:])
            else:
                dq_formatted = dq_clean
            special_remarks_parts.append(f"DIA QLTY-{dq_formatted}")
        special_remarks = ",".join(special_remarks_parts)

        if metal_info["tone_suffix"] == "YG":
            design_inst = "NO RHODIUM"
        elif metal_info["tone_suffix"] == "WG":
            design_inst = "WHITE RHODIUM"
        else:
            design_inst = ""

        item_po_no = order["order_number"]
        if item_po_no and str(item_po_no).isdigit():
            item_po_no = int(item_po_no)

        row = {
            "SrNo": idx,
            "StyleCode": style_code,
            "ItemSize": itemsize,
            "OrderQty": order["qty"],
            "OrderItemPcs": order["qty"],
            "Metal": metal_info["metal"],
            "Tone": metal_info["tone"],
            "ItemPoNo": item_po_no,
            "ItemRefNo": "",
            "StockType": "",
            "MakeType": "",
            "Priority": priority,
            "CustomerProductionInstruction": order["description"],
            "SpecialRemarks": special_remarks,
            "DesignProductionInstruction": design_inst,
            "StampInstruction": order["stamping"] or "",
            "OrderGroup": order_group_val,
            "Certificate": "",
            "SKUNo": order["memo"] or "",
            "Basestoneminwt": "",
            "Basestonemaxwt": "",
            "Basemetalminwt": "",
            "Basemetalmaxwt": "",
            "Productiondeliverydate": "",
            "Expecteddeliverydate": "",
            "BlankColumn": "",
            "SetPrice": "",
            "StoneQuality": "",
            "Date": "",
            "PoDate": order["po_date"] or "",
            "E Del Date": order["due_date"] or "",
        }
        output_rows.append(row)

    return output_rows


def write_output_excel(output_rows, output_path, template_path=None):
    if template_path and os.path.exists(template_path):
        try:
            wb = load_workbook(template_path)
            if "Gati" in wb.sheetnames:
                ws = wb["Gati"]
            else:
                ws = wb.active

            header_cells = {}
            for col_idx in range(1, ws.max_column + 1):
                val = ws.cell(row=1, column=col_idx).value
                if val is not None:
                    header_cells[str(val).strip()] = col_idx

            existing_start = 2
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                has_data = any(cell.value is not None for cell in row)
                if not has_data:
                    break
                existing_start += 1

            if existing_start > 2:
                for row_idx in range(2, existing_start):
                    for col_idx in range(1, ws.max_column + 1):
                        ws.cell(row=row_idx, column=col_idx, value=None)

            for row_idx, row_data in enumerate(output_rows, start=2):
                for col_name, col_idx in header_cells.items():
                    if col_name in row_data:
                        ws.cell(row=row_idx, column=col_idx, value=row_data[col_name])

            #wb.save(output_path)
            return
        except Exception:
            pass

    df = pd.DataFrame(output_rows)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Gati")


def process_bdldhi_file(
    filepath: str,
    output_dir: str,
    recycled: bool = False,
    order_group: str = "",
    priority: str = "",
) -> tuple:
    try:
        reference_df = load_reference_style_map()

        orders = parse_pdf(filepath)
        if not orders:
            return False, None, "Could not extract any line items from the PDF. Please verify the file is a valid SHIMAYRA VPO purchase order.", None

        output_rows = generate_output_rows(
            orders,
            reference_df,
            recycled=recycled,
            order_group=order_group,
            priority=priority,
        )

        base_name = re.sub(r"[^\w\-]", "_", os.path.splitext(os.path.basename(filepath))[0])
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_filename = f"BDLDHI_{base_name}_{timestamp}.xlsx"
        output_path = os.path.join(output_dir, output_filename)

        write_output_excel(output_rows, output_path)

        return True, output_path, None, pd.DataFrame(output_rows)

    except Exception as exc:
        return False, None, str(exc), None